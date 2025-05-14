import os
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from bs4 import BeautifulSoup
import re
from gradio_client import Client
from presidio_analyzer import AnalyzerEngine
from presidio_anonymizer import AnonymizerEngine

# Constants
IN_DIR = '1.IN'
TRANS_DIR = '2.TRANS'
ANON_DIR = '3.ANON'
CHECK_FILE = 'check.xlsx'

# Globals
list_a = []
list_b = []
processed_a = set()
processed_b = set()
lock_a = threading.Lock()
lock_b = threading.Lock()
status_map = {}

# Clients
translator = Client("http://223.194.45.69:7860/")
analyzer = AnalyzerEngine()
anonymizer = AnonymizerEngine()

# Logging
def log_console(msg):
    print(f"{time.strftime('%Y-%m-%d %H:%M:%S')} - {msg}")

# 전처리 함수
def clean_korean_html_text(raw_html):
    soup = BeautifulSoup(str(raw_html), "html.parser")
    text = soup.get_text(separator=" ")
    text = text.replace('\xa0', ' ').replace('&nbsp;', ' ').replace('&quot;', '"')
    text = re.sub(r'\s+', ' ', text).strip()
    text = re.sub(r'https?://\S+|www\.\S+', '', text)
    text = re.sub(r'\b[\w\.-]+@[\w\.-]+\.\w{2,4}\b', '', text)
    text = re.sub(r'\b\d{2,4}[-.\s]??\d{2,4}[-.\s]??\d{4}\b', '', text)
    return text if len(text) > 5 else None

# Function A: 번역 수행
def copy_in_to_trans(filename):
    try:
        if filename in processed_a:
            return
        thread_id = threading.get_ident()
        src = os.path.join(IN_DIR, filename)
        dst = os.path.join(TRANS_DIR, filename)

        wb = openpyxl.load_workbook(src)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        message_idx = headers.index('message') + 1
        translated_idx = len(headers) + 1
        ws.cell(row=1, column=translated_idx, value='translated')

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            raw_html = row[message_idx - 1].value
            clean_text = clean_korean_html_text(raw_html)
            if clean_text:
                try:
                    translated = translator.predict(
                        clean_text,
                        False,
                        "You are a professional English translator. Translate Korean sentences into fluent and grammatically correct English. Respond with only the translated English sentence, without explanation or repetition.",
                        0, 500, 0.5, 0.95, 40, 1.1,
                        api_name="/chat"
                    )
                    row[translated_idx - 1].value = translated
                except Exception as e:
                    row[translated_idx - 1].value = f"ERROR: {e}"

        wb.save(dst)
        with lock_a:
            processed_a.add(filename)
        log_console(f"[A][Thread {thread_id}] Translated and saved: {filename}")
        status_map[filename]['A_status'] = True

    except Exception as e:
        log_console(f"[A][Thread {threading.get_ident()}] ERROR: {filename} - {str(e)}")
        status_map[filename]['A_status'] = False

# Function B: 비식별화 수행
def copy_trans_to_anon(filename):
    try:
        if filename in processed_b:
            return
        thread_id = threading.get_ident()
        src = os.path.join(TRANS_DIR, filename)
        dst = os.path.join(ANON_DIR, filename)

        wb = openpyxl.load_workbook(src)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        translated_idx = headers.index('translated') + 1
        anonymized_idx = len(headers) + 1
        ws.cell(row=1, column=anonymized_idx, value='anonymized')

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            eng_text = row[translated_idx - 1].value
            if eng_text and not str(eng_text).startswith("ERROR"):
                try:
                    results = analyzer.analyze(
                        text=eng_text, 
                        entities=[
                            "PERSON", "EMAIL_ADDRESS", "PHONE_NUMBER", "CREDIT_CARD", "PASSPORT", 
                            "IBAN_CODE", "US_BANK_NUMBER", "NRP", "US_SSN", "IP_ADDRESS", 
                            "DOMAIN_NAME", "MEDICAL_LICENSE", "DRIVER_LICENSE", "MAC_ADDRESS", 
                            "CRYPTO", "IBAN_CODE", "URL", "AWS_ACCESS_KEY", "AWS_SECRET_KEY", 
                            "US_ITIN", "US_DRIVER_LICENSE"
                        ],
                        language='en'
                    )
                    anon_result = anonymizer.anonymize(text=eng_text, analyzer_results=results)
                    row[anonymized_idx - 1].value = anon_result.text
                except Exception as e:
                    row[anonymized_idx - 1].value = f"ERROR: {e}"

        for idx, header in reversed(list(enumerate(headers))):
            if header in ['message', 'translated']:
                ws.delete_cols(idx + 1)
                
        wb.save(dst)
        with lock_b:
            processed_b.add(filename)
        log_console(f"[B][Thread {thread_id}] Anonymized and saved: {filename}")
        status_map[filename]['B_status'] = True

    except Exception as e:
        log_console(f"[B][Thread {threading.get_ident()}] ERROR: {filename} - {str(e)}")
        status_map[filename]['B_status'] = False

# TRANS 폴더 감시
def watch_trans_folder():
    while True:
        time.sleep(2)
        for filename in os.listdir(TRANS_DIR):
            if filename not in processed_b:
                with lock_b:
                    if filename not in list_b:
                        list_b.append(filename)

# 상태 엑셀 저장
def write_status_excel():
    wb = Workbook()
    ws = wb.active
    ws.append(['Filename', 'A_status', 'B_status'])
    for file, stat in status_map.items():
        ws.append([file, stat['A_status'], stat['B_status']])
    wb.save(CHECK_FILE)

# 메인 실행
def main():
    log_console("Scanning IN folder...")
    Path(IN_DIR).mkdir(exist_ok=True)
    Path(TRANS_DIR).mkdir(exist_ok=True)
    Path(ANON_DIR).mkdir(exist_ok=True)

    for filename in os.listdir(IN_DIR):
        list_a.append(filename)
        status_map[filename] = {'A_status': None, 'B_status': None}

    log_console("Started watcher thread for TRANS folder")
    executor_a = ThreadPoolExecutor(max_workers=10)
    executor_b = ThreadPoolExecutor(max_workers=5)

    watcher = threading.Thread(target=watch_trans_folder, daemon=True)
    watcher.start()

    log_console("Phase A: Translation started")
    for filename in list_a:
        executor_a.submit(copy_in_to_trans, filename)

    log_console("Phase B: Waiting for anonymization")
    while True:
        if len(processed_b) == len(list_a):
            break

        with lock_b:
            if list_b:
                chunk = list_b[:2]
                del list_b[:2]
                for file in chunk:
                    executor_b.submit(copy_trans_to_anon, file)

        time.sleep(1)

    executor_a.shutdown(wait=True)
    executor_b.shutdown(wait=True)
    write_status_excel()
    log_console("✅ All processing completed.")

if __name__ == "__main__":
    main()